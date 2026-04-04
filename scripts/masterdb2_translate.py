"""MasterDB2 translation helper: find untranslated items and apply translations.

Usage:
    # Find untranslated items
    python -m scripts.masterdb2_translate scan

    # Find untranslated items and export to JSON
    python -m scripts.masterdb2_translate scan --export /tmp/untranslated.json

    # Apply translations from JSON file
    python -m scripts.masterdb2_translate apply translations.json

    # Show character tone samples for translation reference
    python -m scripts.masterdb2_translate tone [char_code]

JSON format for apply:
    {"FileName": {"原文テキスト": "번역 텍스트", ...}, ...}
"""

import argparse
import json
import os
import sys

import openpyxl
import pandas as pd

from . import paths as _paths


def _read_xlsx(file_path: str) -> pd.DataFrame:
    """Read an xlsx file with the same settings used by the pipeline."""
    df = pd.read_excel(
        file_path,
        na_values="ERROR_NA_VALUE",
        keep_default_na=False,
        na_filter=False,
        engine="openpyxl",
    )
    df.fillna("ERROR_NA_VALUE", inplace=True)
    return df


def scan_untranslated(export_path: str | None = None) -> dict[str, list[dict]]:
    """Scan all masterDB2 xlsx files for untranslated entries.

    Returns dict: {filename: [{keys, id, text}, ...]}
    """
    drive_path = _paths.MASTERDB2_DRIVE_PATH
    result: dict[str, list[dict]] = {}
    total = 0

    for f in sorted(os.listdir(drive_path)):
        if not f.endswith(".xlsx"):
            continue
        name = f[:-5]
        try:
            df = _read_xlsx(os.path.join(drive_path, f))
            if "원문" not in df.columns or "번역" not in df.columns:
                continue
            mask = (df["원문"].astype(str).str.strip() != "") & (
                df["번역"].astype(str).str.strip() == ""
            )
            rows = df[mask]
            if len(rows) == 0:
                continue

            items = []
            for _, row in rows.iterrows():
                keys = "|".join(
                    str(row[c]) for c in df.columns if c.startswith("KEY VALUE")
                )
                items.append(
                    {
                        "keys": keys,
                        "id": str(row.get("ID", "")),
                        "text": str(row["원문"]),
                    }
                )
            result[name] = items
            total += len(items)
            print(f"  {name}: {len(items)}개")
        except Exception as e:
            print(f"  ERROR {name}: {e}", file=sys.stderr)

    print(f"\n총 미번역: {total}개 ({len(result)}개 파일)")

    if export_path:
        with open(export_path, "w", encoding="utf-8") as fp:
            json.dump(result, fp, ensure_ascii=False, indent=2)
        print(f"Exported to {export_path}")

    return result


def apply_translations(trans_path: str) -> None:
    """Apply translations from a JSON file to xlsx files, preserving formatting.

    JSON format: {"FileName": {"原文": "번역", ...}, ...}
    Uses openpyxl to modify cells directly, keeping all formatting intact.
    """
    with open(trans_path, "r", encoding="utf-8") as fp:
        data: dict[str, dict[str, str]] = json.load(fp)

    drive_path = _paths.MASTERDB2_DRIVE_PATH
    total_applied = 0

    for file_name, trans_map in data.items():
        xlsx_path = os.path.join(drive_path, file_name + ".xlsx")
        if not os.path.exists(xlsx_path):
            print(f"  SKIP: {file_name}.xlsx not found")
            continue

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        header_row = [cell.value for cell in ws[1]]
        try:
            orig_col = header_row.index("원문") + 1
            trans_col = header_row.index("번역") + 1
        except ValueError:
            print(f"  SKIP: {file_name} - missing 원문/번역 columns")
            continue

        file_count = 0
        for row in range(2, ws.max_row + 1):
            orig_cell = ws.cell(row=row, column=orig_col)
            trans_cell = ws.cell(row=row, column=trans_col)

            orig_val = str(orig_cell.value) if orig_cell.value is not None else ""
            trans_val = (
                str(trans_cell.value).strip() if trans_cell.value is not None else ""
            )

            if trans_val == "" and orig_val in trans_map:
                trans_cell.value = trans_map[orig_val]
                file_count += 1

        if file_count > 0:
            wb.save(xlsx_path)
            total_applied += file_count
            print(f"  OK: {file_name} - {file_count}개 적용")
        else:
            print(f"  SKIP: {file_name} - no matches")

    print(f"\n총 {total_applied}개 번역 적용")


def show_tone(char_code: str | None = None) -> None:
    """Show existing translated messages per character for tone reference.

    If char_code is given, show only that character. Otherwise show all.
    """
    drive_path = _paths.MASTERDB2_DRIVE_PATH
    xlsx_path = os.path.join(drive_path, "CharacterPushMessage.xlsx")
    if not os.path.exists(xlsx_path):
        print("CharacterPushMessage.xlsx not found")
        return

    df = _read_xlsx(xlsx_path)
    if "KEY VALUE 0" not in df.columns:
        print("Unexpected column structure")
        return

    chars = (
        [char_code]
        if char_code
        else sorted(df["KEY VALUE 0"].astype(str).unique())
    )

    for char in chars:
        if len(char) != 4 or not char.isalpha():
            continue
        mask = (
            (df["KEY VALUE 0"].astype(str) == char)
            & (df["ID"] == "message")
            & (df["번역"].astype(str).str.strip() != "")
        )
        rows = df[mask]
        if len(rows) == 0:
            continue
        print(f"\n=== {char} ({len(rows)}개) ===")
        for _, row in rows.head(6).iterrows():
            typ = str(row.get("KEY VALUE 1", ""))
            orig = str(row["원문"])[:80]
            trans = str(row["번역"])[:80]
            print(f"  [{typ}]")
            print(f"    JP: {orig}")
            print(f"    KR: {trans}")


def main() -> None:
    parser = argparse.ArgumentParser(description="MasterDB2 translation helper")
    sub = parser.add_subparsers(dest="command")

    scan_p = sub.add_parser("scan", help="Scan for untranslated items")
    scan_p.add_argument("--export", help="Export to JSON file")

    apply_p = sub.add_parser("apply", help="Apply translations from JSON")
    apply_p.add_argument("file", help="JSON file with translations")

    tone_p = sub.add_parser("tone", help="Show character tone samples")
    tone_p.add_argument("char", nargs="?", help="Character code (e.g. amao)")

    args = parser.parse_args()

    if args.command == "scan":
        scan_untranslated(args.export)
    elif args.command == "apply":
        apply_translations(args.file)
    elif args.command == "tone":
        show_tone(args.char)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
